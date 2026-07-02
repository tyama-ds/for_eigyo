"""MultiPaperRAG v2 — 複数論文を横断比較するオーケストレーション層。

「広く探す → 論文ごとに深掘り → 突き合わせて比較」を束ねる。

入力:  PDF / Word(.docx) / Excel(.xlsx) を扱える。
図理解: pics=True で、PDF のページ画像を VLM に渡して図・グラフの内容を抽出し、
        比較に利用する（要 pymupdf + 画像対応モデル）。
切替:
  - deep_engine     : "paged"（PagedRAG, 既定・速い）/ "book"（BookRAG, 重いが構造的）
  - locate_strategy : "search"（横断検索, 既定）/ "summary"（要約でLLM選別）/ "all"（全件）
出力:  Comparison.to_df() で pandas DataFrame に変換可能。

接続/プロキシ/埋め込み設定は既存（configure / settings_form）をそのまま使う。

任意依存（未導入なら該当機能のみスキップし案内を出す）:
  Word/Excel 本文 : llama-index-readers-file + docx2txt / openpyxl + pandas
  表抽出          : pdfplumber(PDF) / python-docx(Word) / openpyxl(Excel)
  図理解(pics)    : pymupdf（PDFページ→画像）
  to_df           : pandas
"""

from __future__ import annotations

import base64
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

from .client import complete, get_client
from .config import get_settings
from .pagedrag import PagedRAG

DEFAULT_STORAGE = "./storage/multipaper"
PAPER_EXTS = {".pdf", ".txt", ".md", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".html"}


@dataclass
class Comparison:
    """論文横断の比較結果。"""

    text: str
    papers: list[str] = field(default_factory=list)
    per_paper: dict[str, str] = field(default_factory=dict)
    kind: str = "compare"  # compare / table / figures

    def __str__(self) -> str:
        out = [self.text, "", f"── 対象論文 ({len(self.papers)}) ──"]
        for t in self.papers:
            out.append(f"■ {t}\n{self.per_paper.get(t, '').strip()}")
        return "\n".join(out)

    def to_df(self):
        """論文ごとの結果を pandas DataFrame（index=論文, 列=result）で返す。"""
        try:
            import pandas as pd
        except ImportError as e:
            raise ModuleNotFoundError("to_df には pandas が必要です: pip install pandas") from e
        return pd.DataFrame(
            [{"paper": t, "result": self.per_paper.get(t, "")} for t in self.papers]
        ).set_index("paper")


class MultiPaperRAG:
    """複数論文の横断比較（v2: PDF/Word/Excel、図理解、エンジン/探索の切替、to_df）。"""

    def __init__(
        self,
        storage_dir: str | Path = DEFAULT_STORAGE,
        *,
        top_k: int = 6,
        max_papers: int = 4,
        chunk_size: int = 1024,
        chunk_overlap: int = 128,
        deep_engine: str = "paged",       # "paged" | "book"
        locate_strategy: str = "search",  # "search" | "summary" | "all"
        pics: bool = False,
        vlm_model: str | None = None,     # pics=True 時の画像対応モデル名（省略時は model）
        max_figure_pages: int = 20,
        book_kwargs: dict | None = None,  # deep_engine="book" 時に per-paper BookRAG へ渡す設定
    ):
        if deep_engine not in ("paged", "book"):
            raise ValueError('deep_engine は "paged" か "book"')
        if locate_strategy not in ("search", "summary", "all"):
            raise ValueError('locate_strategy は "search" / "summary" / "all"')
        self.storage_dir = Path(storage_dir)
        self.max_papers = max_papers
        self.deep_engine = deep_engine
        self.locate_strategy = locate_strategy
        self.pics = pics
        self.vlm_model = vlm_model
        self.max_figure_pages = max_figure_pages
        # per-paper BookRAG の調整ノブ（chunk_chars / max_nodes / er_use_llm / max_workers 等）
        self.book_kwargs = book_kwargs or {}
        self._rag = PagedRAG(
            storage_dir=str(self.storage_dir / "vectors"),
            top_k=top_k, chunk_size=chunk_size, chunk_overlap=chunk_overlap,
        )
        self.tables_dir = self.storage_dir / "tables"
        self.figures_dir = self.storage_dir / "figures"
        self._manifest_path = self.storage_dir / "manifest.json"
        self._book_cache: dict[str, object] = {}

    # ---- 取り込み ----------------------------------------------------------

    def add_paper(self, path: str | Path, *, title: str | None = None,
                  extract_tables: bool = True) -> str:
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"ファイルがありません: {path}")
        book_title = self._rag.add_book(path, title=title)
        self._manifest_set(book_title, {"path": str(path.resolve())})
        if extract_tables:
            self._cache_tables(path, book_title)
        if self.pics and path.suffix.lower() == ".pdf":
            self._describe_figures(path, book_title)
        # deep_engine="book" は取り込み時に BookIndex を構築（重い処理と進捗を add 時に出す。
        # 遅延だと初回 compare() で突然重くなるため）。
        if self.deep_engine == "book":
            self._ensure_book(book_title)
        return book_title

    def add_papers(self, docs_dir: str | Path) -> list[str]:
        from .bookindex import progress

        docs_dir = Path(docs_dir)
        files = [f for f in sorted(docs_dir.iterdir())
                 if f.is_file() and f.suffix.lower() in PAPER_EXTS]
        added = []
        for f in progress(files, total=len(files), desc="論文取り込み"):
            print(f"[MultiPaperRAG] 取り込み中: {f.name}")
            added.append(self.add_paper(f))
        return added

    def papers(self) -> list[str]:
        return [b["title"] for b in self._rag.books()]

    # ---- Stage 1: 候補論文の特定（戦略切替） --------------------------------

    def locate(self, question: str, *, max_papers: int | None = None) -> list[str]:
        n = max_papers or self.max_papers
        if self.locate_strategy == "all":
            return self.papers()
        if self.locate_strategy == "summary":
            return self._locate_by_summary(question, n)
        return self._locate_by_search(question, n)

    def _locate_by_search(self, question: str, n: int) -> list[str]:
        ans = self._rag.query(question)
        order: list[str] = []
        for s in ans.sources:
            if s.title and s.title not in order:
                order.append(s.title)
        return order[:n]

    def _locate_by_summary(self, question: str, n: int) -> list[str]:
        manifest = self._manifest()
        summaries = {}
        for t in self.papers():
            summ = manifest.get(t, {}).get("summary")
            if not summ:
                summ = self._make_summary(t)
            summaries[t] = summ
        listing = "\n".join(f"- {t}: {summaries[t]}" for t in summaries)
        prompt = (
            f"質問に答えるのに関連しそうな論文を最大{n}件、関連度順に選んでください。\n"
            'JSON のみ: {"papers": [title, ...]}\n\n'
            f"質問: {question}\n\n論文一覧:\n{listing}"
        )
        from .bookindex import llm_json
        res = llm_json(prompt)
        picked = res.get("papers", []) if isinstance(res, dict) else []
        picked = list(dict.fromkeys(p for p in picked if p in summaries))[:n]  # 重複排除
        return picked or self.papers()[:n]

    def _make_summary(self, title: str) -> str:
        try:
            summ = self._rag.query("この論文の主題・手法・主要な結果を2文で要約して", title=title).text.strip()
        except Exception:  # noqa: BLE001
            summ = ""
        self._manifest_set(title, {"summary": summ})
        return summ

    # ---- compare ------------------------------------------------------------

    def compare(self, question: str, *, papers: list[str] | None = None) -> Comparison:
        cands = papers or self.locate(question) or self.papers()[: self.max_papers]
        if not cands:
            raise RuntimeError("論文が取り込まれていません。add_paper()/add_papers() を実行してください。")
        per_paper: dict[str, str] = {}
        for t in cands:
            ans = self._deep_query(t, question)                # 深掘り（paged/book）
            if self.pics:
                figs = self._load_figures(t)
                if figs:
                    ans += "\n[図の説明]\n" + "\n".join(
                        f"- p.{f['page']}: {f['desc']}" for f in figs[:8]
                    )
            per_paper[t] = ans.strip()
        final = self._synthesize(question, per_paper)
        return Comparison(text=final, papers=cands, per_paper=per_paper, kind="compare")

    def compare_table(self, metric: str, *, papers: list[str] | None = None) -> Comparison:
        cands = papers or self.papers()
        if not cands:
            raise RuntimeError("論文が取り込まれていません。")
        per_paper: dict[str, str] = {}
        for t in cands:
            tables = self._load_tables(t)
            per_paper[t] = (
                self._extract_metric(metric, t, tables) if tables
                else "(表データなし)"
            )
        final = self._synthesize_table(metric, per_paper)
        return Comparison(text=final, papers=cands, per_paper=per_paper, kind="table")

    def compare_figures(self, question: str, *, papers: list[str] | None = None) -> Comparison:
        """pics=True で抽出した図の説明を論文横断で比較する。"""
        if not self.pics:
            raise RuntimeError("compare_figures には pics=True で取り込んだ図の説明が必要です。")
        cands = papers or self.papers()
        per_paper: dict[str, str] = {}
        for t in cands:
            figs = self._load_figures(t)
            per_paper[t] = "\n".join(f"p.{f['page']}: {f['desc']}" for f in figs) or "(図なし)"
        final = self._synthesize(question, per_paper)
        return Comparison(text=final, papers=cands, per_paper=per_paper, kind="figures")

    # ---- 深掘り（エンジン切替） --------------------------------------------

    def _deep_query(self, title: str, question: str) -> str:
        if self.deep_engine == "book":
            book = self._ensure_book(title)
            return book.ask(question).text
        return self._rag.query(question, title=title).text

    def _ensure_book(self, title: str):
        if title in self._book_cache:
            return self._book_cache[title]
        from .bookrag import BookRAG

        book = BookRAG(storage_dir=str(self.storage_dir / "books" / self._safe(title)),
                       **self.book_kwargs)  # 高速化ノブ等を継承（chunk_chars/max_nodes/er_use_llm…）
        info = self._manifest().get(title, {})
        if book.info().get("nodes", 0) == 0:
            src = info.get("path")
            if not src or not Path(src).exists():
                raise RuntimeError(
                    f"論文「{title}」の元ファイルが見つかりません（manifest: {src!r}）。"
                    "ファイルを移動した場合は add_paper で取り込み直してください。"
                )
            book.add_book(src, title=title)  # 初回のみ構築（重い）
        self._book_cache[title] = book
        return book

    # ---- 合成（LLM） --------------------------------------------------------

    def _synthesize(self, question: str, per_paper: dict[str, str]) -> str:
        blocks = "\n\n".join(f"[{t}]\n{ans}" for t, ans in per_paper.items())
        prompt = (
            "以下は複数の論文それぞれから得た情報です。突き合わせて共通点・相違点を明確にし、"
            "質問に答えてください。各主張に論文名を明記し、可能なら Markdown の比較表を添えて"
            f"ください。\n\n質問: {question}\n\n論文別の情報:\n{blocks}"
        )
        return complete(prompt).strip()

    def _extract_metric(self, metric: str, title: str, tables: list[dict]) -> str:
        rendered = self._tables_to_text(tables)
        prompt = (
            f"次は論文「{title}」から抽出した表です。指標「{metric}」に該当する数値を見つけ、"
            "値（条件・データセット名・ページがあれば併記）を簡潔に列挙してください。"
            "該当が無ければ『該当なし』。\n\n" + rendered
        )
        return complete(prompt).strip()

    def _synthesize_table(self, metric: str, per_paper: dict[str, str]) -> str:
        blocks = "\n\n".join(f"[{t}]\n{v}" for t, v in per_paper.items())
        prompt = (
            f"以下は各論文から抽出した「{metric}」の数値です。論文を行、指標/条件を列にした"
            " **Markdown 比較表** を作り、続けて要点を一言述べてください。値が無い論文は空欄に。\n\n"
            + blocks
        )
        return complete(prompt).strip()

    # ---- 表抽出（形式ごとにディスパッチ） -----------------------------------

    def _cache_tables(self, path: Path, title: str) -> None:
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            tables = self._tables_from_pdf(path)
        elif suffix in (".docx", ".doc"):
            tables = self._tables_from_docx(path)
        elif suffix in (".xlsx", ".xls"):
            tables = self._tables_from_excel(path)
        else:
            return
        if tables is None:
            return
        try:
            self.tables_dir.mkdir(parents=True, exist_ok=True)
            (self.tables_dir / f"{self._safe(title)}.json").write_text(
                # default=str: 文字列化を逃れた非 JSON 型が残っても落とさない
                json.dumps(tables, ensure_ascii=False, default=str), encoding="utf-8"
            )
        except Exception as e:  # noqa: BLE001
            print(f"[MultiPaperRAG] 表キャッシュの保存に失敗（表比較なしで続行）: {e}")
            return
        print(f"[MultiPaperRAG] {title}: 表 {len(tables)} 個を抽出")

    def _tables_from_pdf(self, path: Path):
        try:
            import pdfplumber
        except ImportError:
            print("[MultiPaperRAG] PDF 表抽出には pdfplumber が必要: pip install pdfplumber（スキップ）")
            return None
        tables = []
        try:
            with pdfplumber.open(str(path)) as pdf:
                for pno, page in enumerate(pdf.pages, start=1):
                    for tb in (page.extract_tables() or []):
                        if tb:
                            tables.append({"page": pno, "rows": tb})
        except Exception as e:  # noqa: BLE001
            print(f"[MultiPaperRAG] PDF 表抽出に失敗（スキップ）: {e}")
            return None
        return tables

    def _tables_from_docx(self, path: Path):
        try:
            from docx import Document
        except ImportError:
            print("[MultiPaperRAG] Word 表抽出には python-docx が必要: pip install python-docx（スキップ）")
            return None
        tables = []
        try:
            doc = Document(str(path))
            for ti, tb in enumerate(doc.tables, start=1):
                rows = [[cell.text for cell in row.cells] for row in tb.rows]
                if rows:
                    tables.append({"page": ti, "rows": rows})
        except Exception as e:  # noqa: BLE001
            print(f"[MultiPaperRAG] Word 表抽出に失敗（スキップ）: {e}")
            return None
        return tables

    def _tables_from_excel(self, path: Path):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[MultiPaperRAG] Excel 表抽出には openpyxl が必要: pip install openpyxl（スキップ）")
            return None
        tables = []
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            for si, ws in enumerate(wb.worksheets, start=1):
                # datetime 等の非 JSON 型セルは文字列化する（json.dumps でのクラッシュ防止）
                rows = [
                    [("" if c is None else c if isinstance(c, (str, int, float, bool)) else str(c))
                     for c in row]
                    for row in ws.iter_rows(values_only=True)
                ]
                rows = [r for r in rows if any(str(c).strip() for c in r)]
                if rows:
                    tables.append({"page": f"{ws.title}", "rows": rows})
        except Exception as e:  # noqa: BLE001
            print(f"[MultiPaperRAG] Excel 表抽出に失敗（スキップ）: {e}")
            return None
        return tables

    def _load_tables(self, title: str) -> list[dict]:
        p = self.tables_dir / f"{self._safe(title)}.json"
        return json.loads(p.read_text(encoding="utf-8")) if p.exists() else []

    @staticmethod
    def _tables_to_text(tables: list[dict], *, max_tables: int = 20) -> str:
        out = []
        for i, tb in enumerate(tables[:max_tables], start=1):
            lines = [
                " | ".join("" if c is None else str(c) for c in row)
                for row in tb.get("rows", [])
            ]
            out.append(f"# 表{i} ({tb.get('page','?')})\n" + "\n".join(lines))
        return "\n\n".join(out)

    # ---- 図理解（pics=True / VLM） ------------------------------------------

    def _describe_figures(self, path: Path, title: str) -> None:
        try:
            import fitz  # PyMuPDF
        except ImportError:
            print("[MultiPaperRAG] pics=True には pymupdf が必要: pip install pymupdf（図理解をスキップ）")
            return
        figs = []
        try:
            doc = fitz.open(str(path))
            n = min(len(doc), self.max_figure_pages)
            for i in range(n):
                png = doc.load_page(i).get_pixmap(dpi=120).tobytes("png")
                try:
                    desc = self._vlm_describe(png)
                except Exception as e:  # noqa: BLE001
                    print(f"[MultiPaperRAG] VLM 失敗 p.{i+1}（スキップ）: {e}")
                    continue
                if desc and "図なし" not in desc:
                    figs.append({"page": i + 1, "desc": desc.strip()})
            print(f"[MultiPaperRAG] {title}: 図の説明 {len(figs)} ページ分を抽出")
        except Exception as e:  # noqa: BLE001
            print(f"[MultiPaperRAG] 図理解に失敗（スキップ）: {e}")
            return
        self.figures_dir.mkdir(parents=True, exist_ok=True)
        (self.figures_dir / f"{self._safe(title)}.json").write_text(
            json.dumps(figs, ensure_ascii=False), encoding="utf-8"
        )

    def _vlm_describe(self, png_bytes: bytes) -> str:
        s = get_settings()
        model = self.vlm_model or s.model
        b64 = base64.b64encode(png_bytes).decode("ascii")
        resp = get_client().chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": [
                {"type": "text", "text":
                    "この画像中の図・グラフ・チャートの内容と、読み取れる主要な数値・傾向を"
                    "簡潔に説明してください。図やグラフが無ければ「図なし」とだけ答えてください。"},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
            ]}],
            max_tokens=400,
        )
        return resp.choices[0].message.content or ""

    def _load_figures(self, title: str) -> list[dict]:
        p = self.figures_dir / f"{self._safe(title)}.json"
        return json.loads(p.read_text(encoding="utf-8")) if p.exists() else []

    # ---- マニフェスト（title -> path / summary） ----------------------------

    def _manifest(self) -> dict:
        if self._manifest_path.exists():
            return json.loads(self._manifest_path.read_text(encoding="utf-8"))
        return {}

    def _manifest_set(self, title: str, fields: dict) -> None:
        m = self._manifest()
        m.setdefault(title, {}).update(fields)
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self._manifest_path.write_text(json.dumps(m, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _safe(title: str) -> str:
        """タイトルをファイル名に正規化。衝突防止に元タイトルのハッシュを付与する
        （「論文A(2024)」と「論文A（2024）」が同名化して上書きし合う事故を防ぐ）。"""
        import hashlib

        base = re.sub(r"[^\w\-.]+", "_", title)[:100] or "untitled"
        digest = hashlib.md5(title.encode("utf-8")).hexdigest()[:8]
        return f"{base}_{digest}"
